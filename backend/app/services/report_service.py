from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)
from sqlalchemy import case, func, select
from sqlalchemy.orm import Session

from app.models.bill import Bill, BillItem, BillStatus, PaymentMode
from app.models.customer import Customer
from app.models.payment import Payment, PaymentStatus
from app.models.product import Product, ProductVariant
from app.schemas.report import (
    CashBookReport, CashBookRow, DailySalesReport, DailySalesRow,
    EmptyBottleReport, EmptyBottleRow, GstReport, GstSummaryRow,
    OutstandingReport, OutstandingRow, ProductSalesReport, ProductSalesRow,
)

ZERO = Decimal("0")


def daily_sales(db: Session, from_date: date, to_date: date) -> DailySalesReport:
    mode_sum = lambda m: func.sum(case((Bill.payment_mode == m, Bill.amount_paid), else_=ZERO))
    stmt = (
        select(
            Bill.bill_date,
            func.count(Bill.id).label("bills_count"),
            func.sum(Bill.total_amount).label("total_sales"),
            mode_sum(PaymentMode.CASH).label("cash"),
            mode_sum(PaymentMode.CHEQUE).label("cheque"),
            mode_sum(PaymentMode.UPI).label("upi"),
            mode_sum(PaymentMode.CARD).label("card"),
            func.sum(case((Bill.payment_mode == PaymentMode.CREDIT, Bill.total_amount), else_=ZERO)).label("credit"),
        )
        .where(
            Bill.bill_date.between(from_date, to_date),
            Bill.status != BillStatus.CANCELLED,
        )
        .group_by(Bill.bill_date)
        .order_by(Bill.bill_date)
    )
    rows: list[DailySalesRow] = []
    total_sales = ZERO
    total_collected = ZERO
    for r in db.execute(stmt):
        total_sales += r.total_sales or ZERO
        collected = (r.cash or ZERO) + (r.cheque or ZERO) + (r.upi or ZERO) + (r.card or ZERO)
        total_collected += collected
        rows.append(DailySalesRow(
            date=r.bill_date,
            bills_count=r.bills_count,
            total_sales=r.total_sales or ZERO,
            cash_collected=r.cash or ZERO,
            cheque_collected=r.cheque or ZERO,
            upi_collected=r.upi or ZERO,
            card_collected=r.card or ZERO,
            credit_given=r.credit or ZERO,
        ))
    return DailySalesReport(
        from_date=from_date, to_date=to_date, rows=rows,
        grand_total_sales=total_sales, grand_total_collected=total_collected,
    )


def outstanding(db: Session) -> OutstandingReport:
    stmt = (
        select(Customer)
        .where(Customer.is_deleted.is_(False), Customer.current_balance > 0)
        .order_by(Customer.current_balance.desc())
    )
    customers = db.scalars(stmt).all()
    rows = [OutstandingRow(
        customer_id=c.id, customer_name=c.name, mobile=c.mobile, village=c.village,
        current_balance=c.current_balance, current_empty_bottles=c.current_empty_bottles,
    ) for c in customers]
    return OutstandingReport(
        total_customers=len(rows),
        total_outstanding=sum((c.current_balance for c in customers), ZERO),
        rows=rows,
    )


def empty_bottles(db: Session) -> EmptyBottleReport:
    stmt = (
        select(Customer)
        .where(Customer.is_deleted.is_(False), Customer.current_empty_bottles > 0)
        .order_by(Customer.current_empty_bottles.desc())
    )
    customers = db.scalars(stmt).all()
    rows = [EmptyBottleRow(
        customer_id=c.id, customer_name=c.name, mobile=c.mobile, village=c.village,
        empty_bottles=c.current_empty_bottles,
    ) for c in customers]
    return EmptyBottleReport(
        total_empty=sum(c.current_empty_bottles for c in customers),
        rows=rows,
    )


def product_wise_sales(db: Session, from_date: date, to_date: date) -> ProductSalesReport:
    stmt = (
        select(
            BillItem.product_variant_id,
            ProductVariant.name.label("variant_name"),
            Product.name.label("product_name"),
            func.sum(BillItem.quantity).label("qty"),
            func.sum(BillItem.line_total).label("amount"),
        )
        .join(ProductVariant, ProductVariant.id == BillItem.product_variant_id)
        .join(Product, Product.id == ProductVariant.product_id)
        .join(Bill, Bill.id == BillItem.bill_id)
        .where(
            Bill.bill_date.between(from_date, to_date),
            Bill.status != BillStatus.CANCELLED,
        )
        .group_by(BillItem.product_variant_id, ProductVariant.name, Product.name)
        .order_by(func.sum(BillItem.line_total).desc())
    )
    rows = [ProductSalesRow(
        variant_id=r.product_variant_id, variant_name=r.variant_name,
        product_name=r.product_name, qty_sold=int(r.qty or 0),
        total_amount=r.amount or ZERO,
    ) for r in db.execute(stmt)]
    return ProductSalesReport(from_date=from_date, to_date=to_date, rows=rows)


def cash_book(db: Session, from_date: date, to_date: date) -> CashBookReport:
    # cash IN: bills (cash) + payments (cash, cleared)
    bills_cash = (
        select(Bill.bill_date.label("d"), func.sum(Bill.amount_paid).label("amt"))
        .where(
            Bill.bill_date.between(from_date, to_date),
            Bill.status != BillStatus.CANCELLED,
            Bill.payment_mode == PaymentMode.CASH,
        )
        .group_by(Bill.bill_date)
    ).subquery()

    pay_cash = (
        select(Payment.payment_date.label("d"), func.sum(Payment.amount).label("amt"))
        .where(
            Payment.payment_date.between(from_date, to_date),
            Payment.payment_mode == PaymentMode.CASH,
            Payment.status == PaymentStatus.CLEARED,
            Payment.reference_bill_id.is_(None),  # avoid double-count of on-bill payments
        )
        .group_by(Payment.payment_date)
    ).subquery()

    cur = from_date
    rows: list[CashBookRow] = []
    total_in = ZERO
    # build a date-indexed map
    bill_map = {r.d: r.amt or ZERO for r in db.execute(select(bills_cash))}
    pay_map = {r.d: r.amt or ZERO for r in db.execute(select(pay_cash))}
    while cur <= to_date:
        amt_in = (bill_map.get(cur, ZERO) or ZERO) + (pay_map.get(cur, ZERO) or ZERO)
        rows.append(CashBookRow(date=cur, cash_in=amt_in, cash_out=ZERO, net=amt_in))
        total_in += amt_in
        cur += timedelta(days=1)

    return CashBookReport(
        from_date=from_date, to_date=to_date, rows=rows,
        total_in=total_in, total_out=ZERO, net=total_in,
    )


def gst_summary(db: Session, from_date: date, to_date: date) -> GstReport:
    stmt = (
        select(
            BillItem.gst_rate,
            func.sum(BillItem.rate * BillItem.quantity).label("taxable"),
            func.sum(BillItem.gst_amount).label("gst"),
            func.sum(BillItem.line_total).label("total"),
        )
        .join(Bill, Bill.id == BillItem.bill_id)
        .where(
            Bill.bill_date.between(from_date, to_date),
            Bill.status != BillStatus.CANCELLED,
        )
        .group_by(BillItem.gst_rate)
        .order_by(BillItem.gst_rate)
    )
    rows = []
    total_taxable = ZERO
    total_gst = ZERO
    for r in db.execute(stmt):
        rows.append(GstSummaryRow(
            gst_rate=r.gst_rate, taxable_amount=r.taxable or ZERO,
            gst_amount=r.gst or ZERO, total_amount=r.total or ZERO,
        ))
        total_taxable += r.taxable or ZERO
        total_gst += r.gst or ZERO
    return GstReport(
        from_date=from_date, to_date=to_date, rows=rows,
        total_taxable=total_taxable, total_gst=total_gst,
    )


def dashboard(db: Session) -> dict:
    today = date.today()
    today_bills = db.execute(
        select(
            func.count(Bill.id),
            func.coalesce(func.sum(Bill.total_amount), 0),
            func.coalesce(func.sum(case((Bill.payment_mode == PaymentMode.CASH, Bill.amount_paid), else_=0)), 0),
        ).where(Bill.bill_date == today, Bill.status != BillStatus.CANCELLED)
    ).one()
    today_cylinders = db.scalar(
        select(func.coalesce(func.sum(BillItem.quantity), 0))
        .join(Bill, Bill.id == BillItem.bill_id)
        .join(ProductVariant, ProductVariant.id == BillItem.product_variant_id)
        .join(Product, Product.id == ProductVariant.product_id)
        .where(
            Bill.bill_date == today,
            Bill.status != BillStatus.CANCELLED,
            Product.is_returnable.is_(True),
        )
    ) or 0
    outstanding_total = db.scalar(
        select(func.coalesce(func.sum(Customer.current_balance), 0))
        .where(Customer.is_deleted.is_(False), Customer.current_balance > 0)
    ) or 0
    pending_empty = db.scalar(
        select(func.coalesce(func.sum(Customer.current_empty_bottles), 0))
        .where(Customer.is_deleted.is_(False), Customer.current_empty_bottles > 0)
    ) or 0

    return {
        "today_bills_count": today_bills[0] or 0,
        "today_sales_total": today_bills[1] or 0,
        "today_cash_collected": today_bills[2] or 0,
        "today_cylinders_sold": int(today_cylinders),
        "total_outstanding": outstanding_total,
        "total_pending_empty": int(pending_empty),
    }


# ===========================================================================
# Registers — daily and DO-wise rollups for the sidebar "Register" pages.
# ===========================================================================

def _short_serial(bill_no: str | None) -> str:
    """`BILL/26-27/0017` -> `0017`. Anything else returned unchanged."""
    if not bill_no:
        return ""
    return bill_no.rsplit("/", 1)[-1]


def daily_register(
    db: Session,
    from_date: date,
    to_date: date,
    *,
    do_id: int | None = None,
) -> list[dict]:
    """One row per day with the bill# range, count, and total amount."""
    stmt = (
        select(
            Bill.bill_date.label("bill_date"),
            func.min(Bill.bill_number).label("bill_from"),
            func.max(Bill.bill_number).label("bill_to"),
            func.count(Bill.id).label("qty"),
            func.coalesce(func.sum(Bill.total_amount), 0).label("total"),
        )
        .where(
            Bill.bill_date >= from_date,
            Bill.bill_date <= to_date,
            Bill.status == BillStatus.CONFIRMED,
        )
        .group_by(Bill.bill_date)
        .order_by(Bill.bill_date.asc())
    )
    if do_id:
        stmt = stmt.join(Customer, Bill.customer_id == Customer.id).where(
            Customer.do_id == do_id
        )
    rows = db.execute(stmt).all()
    return [
        {
            "date": r.bill_date.isoformat(),
            "bill_from": _short_serial(r.bill_from),
            "bill_to": _short_serial(r.bill_to),
            "qty": int(r.qty),
            "total": str(r.total),
        }
        for r in rows
    ]


def do_register(
    db: Session,
    from_date: date,
    to_date: date,
    *,
    do_id: int | None = None,
) -> list[dict]:
    """One row per (DO, day): includes do_code, do_name + the daily numbers.
    Pass `do_id` to scope the report to a single Distributor Outlet."""
    from app.models.distributor_outlet import DistributorOutlet
    stmt = (
        select(
            DistributorOutlet.id.label("do_id"),
            DistributorOutlet.code.label("do_code"),
            DistributorOutlet.owner_name.label("do_name"),
            DistributorOutlet.location.label("do_location"),
            Bill.bill_date.label("bill_date"),
            func.min(Bill.bill_number).label("bill_from"),
            func.max(Bill.bill_number).label("bill_to"),
            func.count(Bill.id).label("qty"),
            func.coalesce(func.sum(Bill.total_amount), 0).label("total"),
        )
        .join(Customer, Bill.customer_id == Customer.id)
        .join(DistributorOutlet, Customer.do_id == DistributorOutlet.id)
        .where(
            Bill.bill_date >= from_date,
            Bill.bill_date <= to_date,
            Bill.status == BillStatus.CONFIRMED,
        )
        .group_by(
            DistributorOutlet.id,
            DistributorOutlet.code,
            DistributorOutlet.owner_name,
            DistributorOutlet.location,
            Bill.bill_date,
        )
        .order_by(
            DistributorOutlet.code.asc(), Bill.bill_date.asc()
        )
    )
    if do_id:
        stmt = stmt.where(DistributorOutlet.id == do_id)
    rows = db.execute(stmt).all()
    return [
        {
            "do_id": r.do_id,
            "do_code": r.do_code,
            "do_name": r.do_name,
            "do_location": r.do_location,
            "date": r.bill_date.isoformat(),
            "bill_from": _short_serial(r.bill_from),
            "bill_to": _short_serial(r.bill_to),
            "qty": int(r.qty),
            "total": str(r.total),
        }
        for r in rows
    ]


# ---------- Register exports (Excel + PDF) ----------

_HEADER_FILL = PatternFill("solid", fgColor="0F766E")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _fmt_date_str(iso: str) -> str:
    try:
        return datetime.fromisoformat(iso).strftime("%d %b %Y")
    except Exception:
        return iso


def _autofit(ws) -> None:
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, ws.max_column + 1):
        widest = 8
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Skip merged-cell stand-ins (MergedCell rows have no real value).
            v = cell.value
            if v is None:
                continue
            widest = max(widest, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(40, widest + 4)


def daily_register_excel(
    db: Session,
    from_date: date,
    to_date: date,
    *,
    do_id: int | None = None,
) -> bytes:
    rows = daily_register(db, from_date, to_date, do_id=do_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Register"
    ws.append([f"Daily Register · {from_date.strftime('%d %b %Y')} → {to_date.strftime('%d %b %Y')}"])
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=5)
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    headers = ["Date", "Bill # From", "Bill # To", "Qty", "Total (Rs.)"]
    ws.append([])
    ws.append(headers)
    for cell in ws[3]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    total_qty = 0
    total_amt = Decimal("0")
    for r in rows:
        ws.append([
            _fmt_date_str(r["date"]),
            r["bill_from"], r["bill_to"], r["qty"], float(r["total"]),
        ])
        total_qty += r["qty"]
        total_amt += Decimal(r["total"])
    ws.append([])
    ws.append(["TOTAL", "", "", total_qty, float(total_amt)])
    last = ws.max_row
    for cell in ws[last]:
        cell.font = Font(bold=True)
    _autofit(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


_BRAND = colors.HexColor("#0F766E")
_BRAND_LIGHT = colors.HexColor("#CCFBF1")
_BRAND_TINT = colors.HexColor("#E0F2F1")
_TEXT2 = colors.HexColor("#475569")
_BORDER = colors.HexColor("#CBD5E1")
_ZEBRA = colors.HexColor("#F8FAFC")


def _title_para(text: str) -> Paragraph:
    return Paragraph(
        text,
        ParagraphStyle(
            "TitleBig",
            fontName="Helvetica-Bold",
            fontSize=20,
            textColor=_BRAND,
            leading=24,
            spaceAfter=2,
        ),
    )


def _subtitle_para(text: str) -> Paragraph:
    return Paragraph(
        text,
        ParagraphStyle(
            "Sub",
            fontName="Helvetica",
            fontSize=11,
            textColor=_TEXT2,
            leading=14,
            spaceAfter=10,
        ),
    )


def _section_para(text: str) -> Paragraph:
    return Paragraph(
        text,
        ParagraphStyle(
            "Section",
            fontName="Helvetica-Bold",
            fontSize=12,
            textColor=_BRAND,
            leading=16,
            spaceBefore=12,
            spaceAfter=6,
        ),
    )


def _make_table(
    headers: list[str],
    data: list[list],
    *,
    col_widths: list | None = None,
    totals: list | None = None,
    right_align_cols: list[int] | None = None,
    center_align_cols: list[int] | None = None,
) -> Table:
    """Return a styled Table. Body rows alternate with a light stripe; the
    header row has a brand-coloured fill, and optional totals row is bold."""
    body = [headers] + data
    if totals:
        body.append(totals)
    tbl = Table(body, colWidths=col_widths, repeatRows=1)
    cmds = [
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), _BRAND),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10.5),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, 0), 9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 9),
        # Body
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 1), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, _BORDER),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2 if totals else -1),
         [colors.white, _ZEBRA]),
    ]
    for c in right_align_cols or []:
        cmds.append(("ALIGN", (c, 1), (c, -1), "RIGHT"))
        cmds.append(("FONTNAME", (c, 1), (c, -1), "Courier"))
    for c in center_align_cols or []:
        cmds.append(("ALIGN", (c, 1), (c, -1), "CENTER"))
    if totals:
        cmds.extend([
            ("BACKGROUND", (0, -1), (-1, -1), _BRAND_TINT),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, -1), (-1, -1), 10.5),
            ("TEXTCOLOR", (0, -1), (-1, -1), _BRAND),
            ("TOPPADDING", (0, -1), (-1, -1), 9),
            ("BOTTOMPADDING", (0, -1), (-1, -1), 9),
        ])
    tbl.setStyle(TableStyle(cmds))
    return tbl


def _new_doc(buf: BytesIO, *, landscape_mode: bool = False) -> SimpleDocTemplate:
    page = landscape(A4) if landscape_mode else A4
    return SimpleDocTemplate(
        buf, pagesize=page,
        leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=14 * mm, bottomMargin=14 * mm,
        title="SP Gas Register",
    )


def _header_block(title: str, subtitle: str | None = None) -> list:
    block = [_title_para(title)]
    if subtitle:
        block.append(_subtitle_para(subtitle))
    block.append(HRFlowable(
        width="100%", thickness=1.2, color=_BRAND, spaceBefore=2, spaceAfter=12,
    ))
    return block


def daily_register_pdf(
    db: Session,
    from_date: date,
    to_date: date,
    *,
    do_id: int | None = None,
) -> bytes:
    rows = daily_register(db, from_date, to_date, do_id=do_id)
    subtitle_parts = [
        f"{from_date.strftime('%d %b %Y')} → {to_date.strftime('%d %b %Y')}",
        f"{len(rows)} day{'s' if len(rows) != 1 else ''}",
    ]
    buf = BytesIO()
    doc = _new_doc(buf)
    story: list = _header_block(
        "Daily Register", " · ".join(subtitle_parts),
    )

    if not rows:
        story.append(_subtitle_para("No bills in the selected range."))
        doc.build(story)
        return buf.getvalue()

    data = [
        [_fmt_date_str(r["date"]), r["bill_from"], r["bill_to"],
         str(r["qty"]), f"Rs. {Decimal(r['total']):,.2f}"]
        for r in rows
    ]
    total_qty = sum(r["qty"] for r in rows)
    total_amt = sum(Decimal(r["total"]) for r in rows) if rows else Decimal("0")
    totals = ["TOTAL", "", "", str(total_qty), f"Rs. {total_amt:,.2f}"]
    # Wider date column, narrower bill columns, total on the right.
    col_widths = [32 * mm, 28 * mm, 28 * mm, 22 * mm, 38 * mm]
    story.append(_make_table(
        ["Date", "Bill # From", "Bill # To", "Qty", "Total"],
        data,
        col_widths=col_widths,
        totals=totals,
        right_align_cols=[3, 4],
        center_align_cols=[1, 2],
    ))
    doc.build(story)
    return buf.getvalue()


def do_register_excel(
    db: Session,
    from_date: date,
    to_date: date,
    *,
    do_id: int | None = None,
) -> bytes:
    rows = do_register(db, from_date, to_date, do_id=do_id)
    # Exports must read date-first ascending — DO code is secondary tie-break.
    rows.sort(key=lambda r: (r["date"], r["do_code"]))
    wb = Workbook()
    ws = wb.active
    ws.title = "DO Register"
    title = "DO Register"
    if do_id and rows:
        title += f" · {rows[0]['do_code']} / {rows[0]['do_name']}"
    title += (
        f" · {from_date.strftime('%d %b %Y')} → {to_date.strftime('%d %b %Y')}"
    )
    ws.append([title])
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=7)
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    headers = ["DO Code", "DO Name", "Location", "Date",
               "Bill # From", "Bill # To", "Qty", "Total (Rs.)"]
    ws.append([])
    ws.append(headers)
    for cell in ws[3]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    total_qty = 0
    total_amt = Decimal("0")
    for r in rows:
        ws.append([
            r["do_code"], r["do_name"], r.get("do_location") or "",
            _fmt_date_str(r["date"]),
            r["bill_from"], r["bill_to"],
            r["qty"], float(r["total"]),
        ])
        total_qty += r["qty"]
        total_amt += Decimal(r["total"])
    ws.append([])
    ws.append(["TOTAL", "", "", "", "", "", total_qty, float(total_amt)])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    _autofit(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _bills_for_do_pdf(
    db: Session, from_date: date, to_date: date, do_id: int
) -> list[Bill]:
    """Helper used by the DO Register PDF for the bill-by-bill detail section.
    Returns confirmed bills for the DO inside the date range, sorted asc."""
    stmt = (
        select(Bill)
        .join(Customer, Bill.customer_id == Customer.id)
        .where(
            Bill.bill_date >= from_date,
            Bill.bill_date <= to_date,
            Bill.status == BillStatus.CONFIRMED,
            Customer.do_id == do_id,
        )
        .order_by(Bill.bill_date.asc(), Bill.bill_number.asc())
    )
    return list(db.scalars(stmt).all())


def do_register_pdf(
    db: Session,
    from_date: date,
    to_date: date,
    *,
    do_id: int | None = None,
) -> bytes:
    rows = do_register(db, from_date, to_date, do_id=do_id)
    rows.sort(key=lambda r: (r["date"], r["do_code"]))

    # Subtitle line carries the date window + DO scope.
    if do_id and rows:
        title = f"DO Register — {rows[0]['do_code']} / {rows[0]['do_name']}"
        subtitle_lines = [
            (rows[0].get("do_location") or "").strip(),
            f"{from_date.strftime('%d %b %Y')} → {to_date.strftime('%d %b %Y')}",
        ]
        subtitle = " · ".join(s for s in subtitle_lines if s)
    else:
        title = "DO Register — All outlets"
        subtitle = (
            f"{from_date.strftime('%d %b %Y')} → "
            f"{to_date.strftime('%d %b %Y')}"
        )

    buf = BytesIO()
    # Use landscape when we'll show the bill-by-bill detail (more columns).
    landscape_mode = bool(do_id)
    doc = _new_doc(buf, landscape_mode=landscape_mode)
    story: list = _header_block(title, subtitle)

    if not rows:
        story.append(_subtitle_para("No bills in the selected range."))
        doc.build(story)
        return buf.getvalue()

    if do_id:
        # ---------- Single DO: bill-by-bill customer detail only -----------
        bills = _bills_for_do_pdf(db, from_date, to_date, do_id)
        if not bills:
            story.append(_subtitle_para(
                "No bills found for this DO in the selected range."
            ))
        else:
            detail_data = []
            for b in bills:
                cust = b.customer
                short_no = b.bill_number.rsplit("/", 1)[-1]
                name = cust.name if cust else "—"
                # Truncate gracefully so cells don't blow up the row height.
                if len(name) > 32:
                    name = name[:31] + "…"
                mobile = (cust.mobile if cust else "") or "—"
                detail_data.append([
                    f"#{short_no}",
                    b.bill_date.strftime("%d %b %Y"),
                    name,
                    mobile,
                    f"Rs. {Decimal(b.total_amount):,.2f}",
                ])
            d_total = sum(Decimal(b.total_amount) for b in bills)
            d_totals = ["TOTAL", "", "",
                        f"{len(bills)} bills",
                        f"Rs. {d_total:,.2f}"]
            # Landscape A4 usable width ≈ 269mm. Five columns now (no village).
            col_widths = [28 * mm, 38 * mm, 95 * mm, 50 * mm, 58 * mm]
            story.append(_make_table(
                ["Bill #", "Date", "Customer", "Mobile", "Total"],
                detail_data,
                col_widths=col_widths,
                totals=d_totals,
                right_align_cols=[4],
                center_align_cols=[0, 1],
            ))
    else:
        # ---------- All-DOs view: per-(DO, day) rollup ---------------------
        data = [
            [r["do_code"], r["do_name"][:28], _fmt_date_str(r["date"]),
             r["bill_from"], r["bill_to"], str(r["qty"]),
             f"Rs. {Decimal(r['total']):,.2f}"]
            for r in rows
        ]
        total_qty = sum(r["qty"] for r in rows)
        total_amt = sum(Decimal(r["total"]) for r in rows)
        totals = ["TOTAL", "", "", "", "", str(total_qty),
                  f"Rs. {total_amt:,.2f}"]
        col_widths = [22 * mm, 50 * mm, 30 * mm, 24 * mm, 24 * mm,
                      18 * mm, 34 * mm]
        story.append(_make_table(
            ["DO", "Owner", "Date", "Bill # From", "Bill # To",
             "Qty", "Total"],
            data,
            col_widths=col_widths,
            totals=totals,
            right_align_cols=[5, 6],
            center_align_cols=[0, 3, 4],
        ))

    doc.build(story)
    return buf.getvalue()

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Optional

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import Integer, and_, delete, func, or_, select
from sqlalchemy.orm import Session, selectinload

from app.config.settings import settings
from app.models.audit import AuditAction
from app.models.bill import Bill, BillItem, BillStatus, PaymentMode
from app.models.cheque import Cheque, ChequeStatus
from app.models.customer import Customer
from app.models.empty_bottle import EmptyBottleTransaction, EmptyBottleTxnType
from app.models.payment import Payment
from app.models.product import Product, ProductVariant
from app.schemas.bill import BillCreate, BillItemCreate, BillUpdate
from app.utils.audit import write_audit

TWO = Decimal("0.01")


def _fy_prefix(d: date) -> str:
    # Indian FY: Apr-Mar; e.g., 2026-04-20 -> 26-27
    if d.month >= 4:
        start = d.year % 100
        end = (d.year + 1) % 100
    else:
        start = (d.year - 1) % 100
        end = d.year % 100
    return f"{start:02d}-{end:02d}"


def _next_bill_number(db: Session, bill_date: date) -> str:
    """Next bill number = MAX(existing serial in this FY) + 1, or 0001 if none.

    This is the standard accounting behaviour: deleting the latest bill frees
    that number for the next one. Deleting a mid-range bill leaves a hole on
    purpose — gaps preserve the original numbering and can be reconciled
    against the audit log.
    """
    fy = _fy_prefix(bill_date)
    prefix = f"{settings.BILL_CODE_DEFAULT}/{fy}/"
    like_pattern = f"{prefix}%"
    rows = db.scalars(
        select(Bill.bill_number).where(Bill.bill_number.like(like_pattern))
    ).all()
    max_serial = 0
    for r in rows:
        try:
            n = int(r.rsplit("/", 1)[1])
            if n > max_serial:
                max_serial = n
        except (ValueError, IndexError):
            continue
    return f"{prefix}{(max_serial + 1):04d}"


def _load_variant_with_product(db: Session, variant_id: int) -> tuple[ProductVariant, Product]:
    v = db.get(ProductVariant, variant_id)
    if not v or not v.is_active:
        raise HTTPException(status_code=400, detail=f"Variant {variant_id} not found or inactive")
    p = db.get(Product, v.product_id)
    return v, p


def create_bill(db: Session, payload: BillCreate, user_id: int) -> Bill:
    customer = db.get(Customer, payload.customer_id)
    if not customer or customer.is_deleted:
        raise HTTPException(status_code=400, detail="Customer not found")

    bill_date = payload.bill_date or date.today()

    subtotal = Decimal("0")
    gst_total = Decimal("0")
    net_cylinders_issued = 0  # +issued - returned across returnable items
    bill_items: list[BillItem] = []

    for item in payload.items:
        variant, product = _load_variant_with_product(db, item.product_variant_id)
        rate = Decimal(item.rate) if item.rate is not None else variant.unit_price
        gst_rate = Decimal(item.gst_rate) if item.gst_rate is not None else variant.gst_rate
        # rate is GST-inclusive: total = rate * qty; derive gst and base from total
        line_total = (rate * item.quantity).quantize(TWO)
        gst_amount = (line_total * gst_rate / (Decimal(100) + gst_rate)).quantize(TWO)
        line_base = (line_total - gst_amount).quantize(TWO)
        subtotal += line_base
        gst_total += gst_amount

        if product.is_returnable:
            net_cylinders_issued += item.quantity - item.empty_returned

        if variant.stock_quantity is not None and variant.stock_quantity > 0:
            variant.stock_quantity = max(0, variant.stock_quantity - item.quantity)

        bill_items.append(BillItem(
            product_variant_id=variant.id,
            quantity=item.quantity,
            rate=rate,
            empty_returned=item.empty_returned,
            gst_rate=gst_rate,
            gst_amount=gst_amount,
            line_total=line_total,
        ))

    discount = Decimal(payload.discount or 0)
    # subtotal + gst_total already equals sum(line_total) — don't double-add
    total_amount = (subtotal + gst_total - discount).quantize(TWO)
    amount_paid = Decimal(payload.amount_paid or 0).quantize(TWO)
    balance_due = (total_amount - amount_paid).quantize(TWO)

    bill = Bill(
        bill_number=_next_bill_number(db, bill_date),
        bill_date=bill_date,
        customer_id=customer.id,
        subtotal=subtotal.quantize(TWO),
        discount=discount.quantize(TWO),
        gst_amount=gst_total.quantize(TWO),
        total_amount=total_amount,
        amount_paid=amount_paid,
        balance_due=balance_due,
        payment_mode=payload.payment_mode,
        cheque_details=payload.cheque_details.model_dump(mode="json") if payload.cheque_details else None,
        notes=payload.notes,
        status=BillStatus.CONFIRMED,
        created_by_id=user_id,
    )
    bill.items = bill_items
    db.add(bill)
    db.flush()

    # update customer balance
    customer.current_balance = (customer.current_balance + balance_due).quantize(TWO)

    # empty bottle tracking
    if net_cylinders_issued != 0:
        customer.current_empty_bottles += net_cylinders_issued
        txn_type = EmptyBottleTxnType.ISSUED if net_cylinders_issued > 0 else EmptyBottleTxnType.RETURNED
        db.add(EmptyBottleTransaction(
            customer_id=customer.id,
            bill_id=bill.id,
            transaction_type=txn_type,
            quantity=net_cylinders_issued,
            balance_after=customer.current_empty_bottles,
            notes=f"Bill {bill.bill_number}",
            created_by_id=user_id,
        ))

    # cheque register
    if payload.payment_mode == PaymentMode.CHEQUE and payload.cheque_details and amount_paid > 0:
        db.add(Cheque(
            cheque_number=payload.cheque_details.cheque_number,
            bank_name=payload.cheque_details.bank_name,
            branch_name=payload.cheque_details.branch_name,
            cheque_date=payload.cheque_details.cheque_date,
            amount=amount_paid,
            customer_id=customer.id,
            bill_id=bill.id,
            status=ChequeStatus.PENDING,
            created_by_id=user_id,
        ))

    write_audit(db, entity_type="bill", entity_id=bill.id,
                action=AuditAction.CREATE, user_id=user_id,
                changes={"bill_number": bill.bill_number, "total": str(total_amount)})
    db.commit()
    db.refresh(bill)
    return bill


def get_bill(db: Session, bill_id: int) -> Bill:
    bill = db.scalar(
        select(Bill).options(selectinload(Bill.items)).where(Bill.id == bill_id)
    )
    if not bill:
        raise HTTPException(status_code=404, detail="Bill not found")
    return bill


def list_bills(
    db: Session,
    *,
    customer_id: Optional[int] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    status: Optional[BillStatus] = None,
    bill_number_from: Optional[str] = None,
    bill_number_to: Optional[str] = None,
    do_id: Optional[int] = None,
    city: Optional[str] = None,
):
    stmt = select(Bill).order_by(Bill.bill_date.desc(), Bill.id.desc())
    if customer_id:
        stmt = stmt.where(Bill.customer_id == customer_id)
    if from_date:
        stmt = stmt.where(Bill.bill_date >= from_date)
    if to_date:
        stmt = stmt.where(Bill.bill_date <= to_date)
    if status:
        stmt = stmt.where(Bill.status == status)
    # bill_number_from / bill_number_to filter on the **serial** portion only
    # (the trailing digits after the last '/'), so it stays FY-independent —
    # the user typing "5..30" matches BILL/26-27/0005..BILL/26-27/0030 regardless
    # of which date range is also active. Full prefixed strings still work as
    # a fallback string range comparison.
    def _serial_only(v: str) -> Optional[int]:
        s = v.strip().split("/")[-1]
        return int(s) if s.isdigit() else None

    serial_expr = func.cast(
        func.substring(Bill.bill_number, r"[0-9]+$"), Integer
    )
    if bill_number_from:
        n = _serial_only(bill_number_from)
        if n is not None:
            stmt = stmt.where(serial_expr >= n)
        else:
            stmt = stmt.where(Bill.bill_number >= bill_number_from)
    if bill_number_to:
        n = _serial_only(bill_number_to)
        if n is not None:
            stmt = stmt.where(serial_expr <= n)
        else:
            stmt = stmt.where(Bill.bill_number <= bill_number_to)
    if do_id is not None or city:
        stmt = stmt.join(Customer, Bill.customer_id == Customer.id)
        if do_id is not None:
            stmt = stmt.where(Customer.do_id == do_id)
        if city:
            stmt = stmt.where(Customer.city.ilike(f"%{city}%"))
    return stmt


def update_bill(db: Session, bill_id: int, payload: BillUpdate, user_id: int) -> Bill:
    bill = get_bill(db, bill_id)
    if bill.status == BillStatus.CANCELLED:
        raise HTTPException(status_code=400, detail="Cannot edit cancelled bill")

    old_balance = bill.balance_due
    for k, v in payload.model_dump(exclude_unset=True).items():
        if k == "cheque_details" and v is not None:
            # keep as plain dict
            bill.cheque_details = payload.cheque_details.model_dump(mode="json")
        else:
            setattr(bill, k, v)

    # recompute balance if discount or amount_paid changed
    bill.balance_due = (bill.total_amount - Decimal(bill.amount_paid or 0)).quantize(TWO)

    # adjust customer balance by delta
    diff = bill.balance_due - old_balance
    if diff:
        customer = db.get(Customer, bill.customer_id)
        customer.current_balance = (customer.current_balance + diff).quantize(TWO)

    write_audit(db, entity_type="bill", entity_id=bill.id,
                action=AuditAction.UPDATE, user_id=user_id)
    db.commit()
    db.refresh(bill)
    return bill


def cancel_bill(db: Session, bill_id: int, user_id: int) -> Bill:
    bill = get_bill(db, bill_id)
    if bill.status == BillStatus.CANCELLED:
        raise HTTPException(status_code=400, detail="Bill already cancelled")

    customer = db.get(Customer, bill.customer_id)
    # reverse balance
    customer.current_balance = (customer.current_balance - bill.balance_due).quantize(TWO)

    # reverse empty bottle transactions for this bill
    reversals = db.scalars(
        select(EmptyBottleTransaction).where(EmptyBottleTransaction.bill_id == bill.id)
    ).all()
    for tx in reversals:
        customer.current_empty_bottles -= tx.quantity
    if reversals:
        db.add(EmptyBottleTransaction(
            customer_id=customer.id,
            bill_id=bill.id,
            transaction_type=EmptyBottleTxnType.ADJUSTMENT,
            quantity=-sum(tx.quantity for tx in reversals),
            balance_after=customer.current_empty_bottles,
            notes=f"Cancel bill {bill.bill_number}",
            created_by_id=user_id,
        ))

    # restore stock
    for item in bill.items:
        v = db.get(ProductVariant, item.product_variant_id)
        if v:
            v.stock_quantity += item.quantity

    # cancel associated cheque(s)
    for cheque in db.scalars(select(Cheque).where(Cheque.bill_id == bill.id)).all():
        if cheque.status == ChequeStatus.PENDING:
            cheque.status = ChequeStatus.CANCELLED

    bill.status = BillStatus.CANCELLED
    write_audit(db, entity_type="bill", entity_id=bill.id,
                action=AuditAction.CANCEL, user_id=user_id)
    db.commit()
    db.refresh(bill)
    return bill


def delete_bill_hard(db: Session, bill_id: int, user_id: int) -> dict:
    """Reverse a bill's side-effects (like cancel) AND remove the row from DB
    so its bill_number is free again — the next created bill will reuse the
    smallest gap, including this one.

    Side-effects reversed: customer balance, empty-bottle ledger, stock,
    bill-tied cheques and payments. Audit row is written with the bill # so
    the deletion is traceable even after the row is gone.
    """
    bill = get_bill(db, bill_id)
    bill_number = bill.bill_number
    customer = db.get(Customer, bill.customer_id)

    # Only reverse customer balance / empties if the bill was ACTIVE (not
    # already cancelled — cancel_bill would have reversed them once already).
    if bill.status != BillStatus.CANCELLED:
        customer.current_balance = (
            customer.current_balance - bill.balance_due
        ).quantize(TWO)

        empties = db.scalars(
            select(EmptyBottleTransaction).where(
                EmptyBottleTransaction.bill_id == bill.id
            )
        ).all()
        for tx in empties:
            customer.current_empty_bottles -= tx.quantity

        # Restore stock
        for item in bill.items:
            v = db.get(ProductVariant, item.product_variant_id)
            if v and v.stock_quantity is not None:
                v.stock_quantity += item.quantity

    # Drop bill-tied cheques and payments (no audit cancel — full delete).
    db.execute(
        delete(Cheque).where(Cheque.bill_id == bill.id)
    )
    db.execute(
        delete(Payment).where(Payment.reference_bill_id == bill.id)
    )
    # Drop empty-bottle txns linked to this bill so they don't dangle.
    db.execute(
        delete(EmptyBottleTransaction).where(
            EmptyBottleTransaction.bill_id == bill.id
        )
    )

    # Audit BEFORE deleting the row.
    write_audit(
        db, entity_type="bill", entity_id=bill.id,
        action=AuditAction.DELETE, user_id=user_id,
        changes={"bill_number": bill_number, "hard_delete": True},
    )

    db.delete(bill)  # bill_items cascade
    db.commit()
    return {"bill_number": bill_number}


def bulk_delete_bills(
    db: Session, bill_ids: list[int], user_id: int
) -> dict:
    """Hard-delete many bills; each one reverses its side-effects (customer
    balance, empties, stock, cheques, payments) and removes the row so the
    bill # is free again. Missing ids are skipped silently."""
    if not bill_ids:
        return {"deleted": 0, "skipped": 0, "bill_numbers": []}

    deleted_numbers: list[str] = []
    skipped = 0
    requested = set(bill_ids)
    for bid in bill_ids:
        try:
            bill = db.scalar(
                select(Bill)
                .options(selectinload(Bill.items))
                .where(Bill.id == bid)
            )
            if not bill:
                skipped += 1
                continue
            bill_number = bill.bill_number
            customer = db.get(Customer, bill.customer_id)

            if bill.status != BillStatus.CANCELLED:
                customer.current_balance = (
                    customer.current_balance - bill.balance_due
                ).quantize(TWO)
                empties = db.scalars(
                    select(EmptyBottleTransaction).where(
                        EmptyBottleTransaction.bill_id == bill.id
                    )
                ).all()
                for tx in empties:
                    customer.current_empty_bottles -= tx.quantity
                for item in bill.items:
                    v = db.get(ProductVariant, item.product_variant_id)
                    if v and v.stock_quantity is not None:
                        v.stock_quantity += item.quantity

            db.execute(delete(Cheque).where(Cheque.bill_id == bill.id))
            db.execute(
                delete(Payment).where(Payment.reference_bill_id == bill.id)
            )
            db.execute(
                delete(EmptyBottleTransaction).where(
                    EmptyBottleTransaction.bill_id == bill.id
                )
            )
            db.delete(bill)
            db.flush()
            deleted_numbers.append(bill_number)
        except Exception:
            db.rollback()
            skipped += 1

    write_audit(
        db, entity_type="bill", entity_id=None,
        action=AuditAction.DELETE, user_id=user_id,
        changes={
            "bulk": True,
            "deleted_count": len(deleted_numbers),
            "deleted_numbers": deleted_numbers,
            "requested": len(requested),
        },
    )
    db.commit()
    return {
        "deleted": len(deleted_numbers),
        "skipped": skipped,
        "bill_numbers": deleted_numbers,
    }


def customer_ledger(db: Session, customer_id: int):
    from app.models.payment import Payment
    customer = db.get(Customer, customer_id)
    if not customer or customer.is_deleted:
        raise HTTPException(status_code=404, detail="Customer not found")

    entries = []
    running = Decimal(customer.opening_balance or 0)
    entries.append({
        "date": customer.registration_date, "type": "opening",
        "reference": "OPENING", "debit": running, "credit": Decimal("0"),
        "balance": running, "notes": "Opening balance",
    })

    bills = db.scalars(
        select(Bill).where(Bill.customer_id == customer_id, Bill.status != BillStatus.CANCELLED)
        .order_by(Bill.bill_date, Bill.id)
    ).all()
    payments = db.scalars(
        select(Payment).where(Payment.customer_id == customer_id)
        .order_by(Payment.payment_date, Payment.id)
    ).all()

    # merge and order by date
    events = [("bill", b) for b in bills] + [("payment", p) for p in payments]
    events.sort(key=lambda e: (e[1].bill_date if e[0] == "bill" else e[1].payment_date, e[1].id))

    for kind, ev in events:
        if kind == "bill":
            running += ev.balance_due
            entries.append({
                "date": ev.bill_date, "type": "bill", "reference": ev.bill_number,
                "debit": ev.total_amount, "credit": ev.amount_paid,
                "balance": running, "notes": ev.notes,
            })
        else:
            running -= ev.amount
            entries.append({
                "date": ev.payment_date, "type": "payment", "reference": ev.payment_number,
                "debit": Decimal("0"), "credit": ev.amount,
                "balance": running, "notes": ev.notes,
            })

    return {
        "customer_id": customer.id,
        "customer_name": customer.name,
        "mobile": customer.mobile,
        "opening_balance": customer.opening_balance,
        "entries": entries,
        "closing_balance": running,
    }


# ===========================================================================
# Reset / hard-delete all bills — admin "start fresh" action.
# ===========================================================================
def reset_all_bills(db: Session, user_id: int) -> dict:
    """Delete every bill (and bill_items via cascade), reverse the side-effects
    on customers/empties/cheques/payments, so the next bill starts at 0001.

    Steps:
      1. Drop all bill-linked Cheque rows (no payment_id linkage left).
      2. Drop all bill-linked Payment rows (Cheques may have already removed).
      3. Drop all bill-linked EmptyBottleTransaction rows.
      4. Delete all Bill rows (bill_items cascade automatically).
      5. Reset every customer's current_balance and current_empty_bottles to
         their opening values (any standalone Payment recomputed below).
      6. Re-apply remaining standalone payments / non-bill empty txns.
    """
    bill_count = db.scalar(select(func.count()).select_from(Bill)) or 0

    # 1. Cheques tied to bills (bill-linked; standalone cheques have NULL FK)
    db.execute(delete(Cheque).where(Cheque.bill_id.is_not(None)))
    # 2. Payments tied to bills — Payment uses reference_bill_id (FK SET NULL).
    #    Delete them so the bill totally disappears from customer ledger.
    db.execute(delete(Payment).where(Payment.reference_bill_id.is_not(None)))
    # 3. Empty bottle txns tied to bills
    db.execute(delete(EmptyBottleTransaction).where(
        EmptyBottleTransaction.bill_id.is_not(None)))
    # 4. Bills (bill_items cascade)
    db.execute(delete(Bill))
    db.flush()

    # 5+6. Recompute every customer's balance/empty count from opening +
    # surviving (truly standalone) payments and empty txns.
    customers = list(db.scalars(
        select(Customer).where(Customer.is_deleted.is_(False))
    ).all())
    for cust in customers:
        paid_total = db.scalar(
            select(func.coalesce(func.sum(Payment.amount), 0)).where(
                Payment.customer_id == cust.id,
            )
        ) or 0
        cust.current_balance = (
            Decimal(cust.opening_balance or 0) - Decimal(paid_total)
        ).quantize(TWO)

        # Empty bottles: opening + remaining (non-bill, non-opening) empty txns.
        empties_delta = db.scalar(
            select(func.coalesce(func.sum(EmptyBottleTransaction.quantity), 0))
            .where(
                EmptyBottleTransaction.customer_id == cust.id,
                EmptyBottleTransaction.bill_id.is_(None),
                EmptyBottleTransaction.transaction_type
                != EmptyBottleTxnType.OPENING,
            )
        ) or 0
        cust.current_empty_bottles = (
            (cust.opening_empty_bottles or 0) + int(empties_delta)
        )

    write_audit(
        db, entity_type="bill", entity_id=None,
        action=AuditAction.DELETE, user_id=user_id,
        changes={"reset": True, "bills_deleted": int(bill_count)},
    )
    db.commit()
    return {
        "bills_deleted": int(bill_count),
        "customers_reset": len(customers),
    }


# ===========================================================================
# Bulk-import bills from Excel.
# Headers (case-insensitive, any subset; only Mobile + Date are required):
#   Mobile  | Date  | Variant | Qty | Amount_Paid | Payment_Mode |
#   Empty_Returned | Discount | Notes
# ===========================================================================
_DATE_FORMATS = ("%d.%m.%Y", "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d.%m.%y")
_VALID_MODES = {m.value for m in PaymentMode}


def _parse_bill_date(raw) -> Optional[date]:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = str(raw).strip()
    if not s:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unrecognised date '{s}' (expected DD.MM.YYYY)")


def _resolve_variant(
    db: Session, raw, fallback: ProductVariant
) -> ProductVariant:
    """Match by exact id, exact SKU, or case-insensitive name partial."""
    if raw is None or str(raw).strip() == "":
        return fallback
    s = str(raw).strip()
    if s.isdigit():
        v = db.get(ProductVariant, int(s))
        if v and v.is_active:
            return v
    # exact SKU
    v = db.scalar(
        select(ProductVariant).where(
            func.lower(ProductVariant.sku_code) == s.lower(),
            ProductVariant.is_active.is_(True),
        )
    )
    if v:
        return v
    # name partial
    v = db.scalar(
        select(ProductVariant).where(
            ProductVariant.name.ilike(f"%{s}%"),
            ProductVariant.is_active.is_(True),
        )
    )
    if v:
        return v
    raise ValueError(f"Variant '{s}' not found")


def import_bills_from_excel(
    db: Session, file_bytes: bytes, user_id: int
) -> dict:
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    # Lowercase header lookup so 'Mobile', 'mobile', 'MOBILE' all work.
    raw_headers = [str(c).strip() if c else "" for c in rows[0]]
    headers = [h.lower() for h in raw_headers]

    def col(row_dict: dict, *keys) -> Any:
        for k in keys:
            if k in row_dict:
                return row_dict[k]
        return None

    # Default variant — first active. Used when row leaves Variant blank.
    default_variant = db.scalar(
        select(ProductVariant)
        .where(ProductVariant.is_active.is_(True))
        .order_by(ProductVariant.id.asc())
    )
    if not default_variant:
        raise HTTPException(
            status_code=400,
            detail="No active product variant exists. Add one before importing bills.",
        )

    imported = 0
    errors: list[dict] = []
    for idx, raw in enumerate(rows[1:], start=2):
        if not any(raw):
            continue  # skip blank lines
        row = {
            headers[i]: (raw[i] if i < len(raw) else None)
            for i in range(len(headers))
        }
        try:
            mobile = str(col(row, "mobile") or "").strip()
            if not mobile:
                raise ValueError("Mobile is required")
            cust = db.scalar(
                select(Customer).where(
                    Customer.mobile == mobile,
                    Customer.is_deleted.is_(False),
                )
            )
            if not cust:
                raise ValueError(f"Customer with mobile {mobile} not found")

            bill_date = _parse_bill_date(col(row, "date", "bill_date"))
            if not bill_date:
                raise ValueError("Date is required")

            variant = _resolve_variant(
                db, col(row, "variant", "variant_code", "sku"), default_variant
            )

            qty_raw = col(row, "qty", "quantity")
            qty = int(qty_raw) if qty_raw not in (None, "") else 1
            if qty <= 0:
                raise ValueError("Qty must be > 0")

            empty_ret_raw = col(row, "empty_returned", "empty")
            empty_ret = (
                int(empty_ret_raw) if empty_ret_raw not in (None, "") else 0
            )

            amt_raw = col(row, "amount_paid", "paid")
            amt_paid = (
                Decimal(str(amt_raw)) if amt_raw not in (None, "") else Decimal("0")
            )

            disc_raw = col(row, "discount")
            discount = (
                Decimal(str(disc_raw))
                if disc_raw not in (None, "") else Decimal("0")
            )

            mode_raw = (
                str(col(row, "payment_mode", "mode") or "cash").strip().lower()
            )
            if mode_raw not in _VALID_MODES:
                raise ValueError(
                    f"payment_mode '{mode_raw}' invalid (use: {sorted(_VALID_MODES)})"
                )
            mode = PaymentMode(mode_raw)

            notes = col(row, "notes")
            notes = str(notes).strip() if notes else None

            payload = BillCreate(
                bill_date=bill_date,
                customer_id=cust.id,
                items=[
                    BillItemCreate(
                        product_variant_id=variant.id,
                        quantity=qty,
                        empty_returned=empty_ret,
                    )
                ],
                discount=discount,
                amount_paid=amt_paid,
                payment_mode=mode,
                notes=notes,
            )
            create_bill(db, payload, user_id)
            imported += 1
        except Exception as e:
            errors.append({"row": idx, "error": str(e), "data": dict(row)})

    write_audit(
        db, entity_type="bill", entity_id=None,
        action=AuditAction.IMPORT, user_id=user_id,
        changes={"imported": imported, "errors": len(errors)},
    )
    return {"imported": imported, "errors": errors}

from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Optional

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from sqlalchemy import and_, or_, select
from sqlalchemy.orm import Session

from app.models.audit import AuditAction
from app.models.customer import Customer, CustomerStatus, CustomerType
from app.models.distributor_outlet import DistributorOutlet
from app.models.empty_bottle import EmptyBottleTransaction, EmptyBottleTxnType
from app.schemas.customer import (
    CustomerCreate, CustomerImportError, CustomerImportResult, CustomerUpdate,
)
from app.utils.audit import write_audit


def _check_mobile_unique(
    db: Session, mobile: str, exclude_id: Optional[int] = None
) -> None:
    stmt = select(Customer).where(
        Customer.mobile == mobile,
        Customer.is_deleted.is_(False),
    )
    if exclude_id:
        stmt = stmt.where(Customer.id != exclude_id)
    if db.scalar(stmt):
        raise HTTPException(
            status_code=400,
            detail=f"Customer already exists with mobile {mobile}",
        )


def _check_consumer_number_unique(
    db: Session, consumer_number: Optional[str], exclude_id: Optional[int] = None
) -> None:
    if not consumer_number:
        return
    stmt = select(Customer).where(
        Customer.consumer_number == consumer_number,
        Customer.is_deleted.is_(False),
    )
    if exclude_id:
        stmt = stmt.where(Customer.id != exclude_id)
    if db.scalar(stmt):
        raise HTTPException(
            status_code=400,
            detail=f"Consumer number '{consumer_number}' already exists",
        )


def _validate_do(db: Session, do_id: int) -> DistributorOutlet:
    do = db.get(DistributorOutlet, do_id)
    if not do or do.is_deleted:
        raise HTTPException(status_code=400, detail="Distributor outlet not found")
    if not do.is_active:
        raise HTTPException(status_code=400, detail="Distributor outlet is inactive")
    return do


def create_customer(db: Session, payload: CustomerCreate, user_id: int) -> Customer:
    _check_mobile_unique(db, payload.mobile)
    _check_consumer_number_unique(db, payload.consumer_number)
    _validate_do(db, payload.do_id)

    cust = Customer(
        consumer_number=payload.consumer_number,
        do_id=payload.do_id,
        name=payload.name,
        mobile=payload.mobile,
        alternate_mobile=payload.alternate_mobile,
        village=payload.village,
        city=payload.city,
        district=payload.district,
        state=payload.state or "Gujarat",
        pincode=payload.pincode,
        full_address=payload.full_address,
        customer_type=payload.customer_type,
        aadhaar_number=payload.aadhaar_number,
        email=payload.email,
        date_of_birth=payload.date_of_birth,
        notes=payload.notes,
        registration_date=payload.registration_date or date.today(),
        status=payload.status,
        opening_balance=payload.opening_balance,
        opening_empty_bottles=payload.opening_empty_bottles,
        current_balance=payload.opening_balance,
        current_empty_bottles=payload.opening_empty_bottles,
        created_by_id=user_id,
    )
    db.add(cust)
    db.flush()

    if payload.opening_empty_bottles:
        db.add(EmptyBottleTransaction(
            customer_id=cust.id,
            transaction_type=EmptyBottleTxnType.OPENING,
            quantity=payload.opening_empty_bottles,
            balance_after=payload.opening_empty_bottles,
            notes="Opening balance",
            created_by_id=user_id,
        ))

    write_audit(db, entity_type="customer", entity_id=cust.id,
                action=AuditAction.CREATE, user_id=user_id,
                changes={"name": cust.name, "mobile": cust.mobile})
    db.commit()
    db.refresh(cust)
    return cust


def update_customer(db: Session, customer_id: int, payload: CustomerUpdate, user_id: int) -> Customer:
    cust = get_customer(db, customer_id)
    data = payload.model_dump(exclude_unset=True)

    if "mobile" in data and data["mobile"] != cust.mobile:
        _check_mobile_unique(db, data["mobile"], exclude_id=cust.id)

    if "consumer_number" in data and data["consumer_number"] != cust.consumer_number:
        _check_consumer_number_unique(db, data["consumer_number"], exclude_id=cust.id)

    if "do_id" in data and data["do_id"] != cust.do_id:
        _validate_do(db, data["do_id"])

    changes = {}
    for k, v in data.items():
        old = getattr(cust, k)
        if old != v:
            changes[k] = {"old": str(old) if old is not None else None, "new": str(v) if v is not None else None}
            setattr(cust, k, v)

    if changes:
        write_audit(db, entity_type="customer", entity_id=cust.id,
                    action=AuditAction.UPDATE, user_id=user_id, changes=changes)
    db.commit()
    db.refresh(cust)
    return cust


def get_customer(db: Session, customer_id: int) -> Customer:
    cust = db.get(Customer, customer_id)
    if not cust or cust.is_deleted:
        raise HTTPException(status_code=404, detail="Customer not found")
    return cust


def soft_delete_customer(db: Session, customer_id: int, user_id: int) -> None:
    cust = get_customer(db, customer_id)
    cust.is_deleted = True
    cust.status = CustomerStatus.INACTIVE
    write_audit(db, entity_type="customer", entity_id=cust.id,
                action=AuditAction.DELETE, user_id=user_id)
    db.commit()


def set_customer_active(db: Session, customer_id: int, active: bool, user_id: int) -> Customer:
    cust = get_customer(db, customer_id)
    new_status = CustomerStatus.ACTIVE if active else CustomerStatus.INACTIVE
    if cust.status == new_status:
        return cust
    cust.status = new_status
    write_audit(db, entity_type="customer", entity_id=cust.id,
                action=AuditAction.UPDATE, user_id=user_id,
                changes={"status": new_status.value})
    db.commit()
    db.refresh(cust)
    return cust


def list_customers(
    db: Session,
    *,
    q: Optional[str] = None,
    customer_type: Optional[CustomerType] = None,
    status: Optional[CustomerStatus] = None,
    village: Optional[str] = None,
    include_deleted: bool = False,
):
    stmt = select(Customer)
    if not include_deleted:
        stmt = stmt.where(Customer.is_deleted.is_(False))
    if q:
        like = f"%{q}%"
        stmt = stmt.where(or_(
            Customer.name.ilike(like),
            Customer.mobile.ilike(like),
            Customer.village.ilike(like),
            Customer.city.ilike(like),
        ))
    if customer_type:
        stmt = stmt.where(Customer.customer_type == customer_type)
    if status:
        stmt = stmt.where(Customer.status == status)
    if village:
        stmt = stmt.where(Customer.village.ilike(f"%{village}%"))
    stmt = stmt.order_by(Customer.name.asc())
    return stmt


def search_customers(db: Session, q: str, limit: int = 20) -> list[Customer]:
    if not q or len(q.strip()) < 2:
        return []
    like = f"%{q.strip()}%"
    stmt = (
        select(Customer)
        .where(
            Customer.is_deleted.is_(False),
            or_(
                Customer.mobile.ilike(like),
                and_(Customer.name.ilike(like)),
                Customer.village.ilike(like),
            ),
        )
        .order_by(Customer.name.asc())
        .limit(limit)
    )
    return list(db.scalars(stmt).all())


# --------- Excel import/export ---------
IMPORT_HEADERS = ["Name", "Mobile", "DO", "Village", "City", "Type", "Opening_Balance", "Opening_Bottles"]


def import_customers_from_excel(
    db: Session, file_bytes: bytes, user_id: int
) -> CustomerImportResult:
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    headers = [str(c).strip() if c else "" for c in rows[0]]
    errors: list[CustomerImportError] = []
    imported = 0
    skipped = 0

    # Build a code → DO map for lookups; first active DO is the fallback when
    # the row leaves the DO column blank.
    active_dos = list(db.scalars(
        select(DistributorOutlet).where(
            DistributorOutlet.is_deleted.is_(False),
            DistributorOutlet.is_active.is_(True),
        ).order_by(DistributorOutlet.id.asc())
    ).all())
    if not active_dos:
        raise HTTPException(status_code=400, detail="No active Distributor Outlet exists. Create one before importing.")
    do_by_code = {d.code.upper(): d for d in active_dos}
    fallback_do = active_dos[0]

    for idx, raw in enumerate(rows[1:], start=2):
        row = {headers[i]: (raw[i] if i < len(raw) else None) for i in range(len(headers))}
        try:
            name = str(row.get("Name", "") or "").strip()
            mobile = str(row.get("Mobile", "") or "").strip()
            village = str(row.get("Village", "") or "").strip() or None
            city = str(row.get("City", "") or "").strip()
            consumer_number = str(row.get("Consumer_Number", "") or "").strip() or None
            if not name or not mobile:
                raise ValueError("Name and Mobile are required")
            if not mobile.isdigit() or len(mobile) < 10:
                raise ValueError("Mobile must be 10+ digits")

            # DO column: accept "DO" or "DO_Code" header. Empty → fallback DO.
            do_raw = (row.get("DO") or row.get("DO_Code") or "")
            do_code = str(do_raw).strip().upper()
            if do_code:
                do = do_by_code.get(do_code)
                if not do:
                    raise ValueError(
                        f"DO code '{do_code}' not found (active codes: "
                        f"{', '.join(sorted(do_by_code.keys()))})"
                    )
            else:
                do = fallback_do

            ctype_raw = (str(row.get("Type", "") or "domestic")).strip().lower()
            ctype = CustomerType.COMMERCIAL if ctype_raw.startswith("c") else CustomerType.DOMESTIC
            op_bal = Decimal(str(row.get("Opening_Balance", 0) or 0))
            op_bot = int(row.get("Opening_Bottles", 0) or 0)

            # skip if duplicate mobile OR duplicate consumer_number
            exists = db.scalar(select(Customer).where(
                Customer.mobile == mobile,
                Customer.is_deleted.is_(False),
            ))
            if exists:
                skipped += 1
                continue
            if consumer_number and db.scalar(select(Customer).where(
                Customer.consumer_number == consumer_number,
                Customer.is_deleted.is_(False),
            )):
                skipped += 1
                continue

            cust = Customer(
                consumer_number=consumer_number,
                do_id=do.id,
                name=name, mobile=mobile, village=village, city=city or village or "—",
                customer_type=ctype, registration_date=date.today(),
                opening_balance=op_bal, opening_empty_bottles=op_bot,
                current_balance=op_bal, current_empty_bottles=op_bot,
                created_by_id=user_id,
            )
            db.add(cust)
            db.flush()
            if op_bot:
                db.add(EmptyBottleTransaction(
                    customer_id=cust.id,
                    transaction_type=EmptyBottleTxnType.OPENING,
                    quantity=op_bot, balance_after=op_bot,
                    notes="Opening balance (import)", created_by_id=user_id,
                ))
            imported += 1
        except Exception as e:
            errors.append(CustomerImportError(row=idx, error=str(e), data=row))

    write_audit(db, entity_type="customer", entity_id=None,
                action=AuditAction.IMPORT, user_id=user_id,
                changes={"imported": imported, "skipped": skipped, "errors": len(errors)})
    db.commit()
    return CustomerImportResult(imported=imported, skipped=skipped, errors=errors)


def export_customers_to_excel(db: Session) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    headers = [
        "ID", "Consumer Number", "DO Code", "DO Owner", "Name", "Mobile", "Alt Mobile",
        "Village", "City", "District", "State", "Pincode", "Type", "Status",
        "Current Balance", "Empty Bottles", "Registration Date",
    ]
    ws.append(headers)
    for cust in db.scalars(select(Customer).where(Customer.is_deleted.is_(False)).order_by(Customer.name)):
        do = cust.distributor_outlet
        ws.append([
            cust.id, cust.consumer_number,
            do.code if do else "", do.owner_name if do else "",
            cust.name, cust.mobile, cust.alternate_mobile,
            cust.village, cust.city, cust.district, cust.state, cust.pincode,
            cust.customer_type.value, cust.status.value,
            float(cust.current_balance), cust.current_empty_bottles,
            cust.registration_date.isoformat() if cust.registration_date else "",
        ])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
